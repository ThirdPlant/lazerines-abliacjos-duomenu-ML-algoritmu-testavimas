import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.compose import ColumnTransformer
from sklearn.gaussian_process import GaussianProcessRegressor
from sklearn.gaussian_process.kernels import ConstantKernel, Matern, WhiteKernel
from sklearn.preprocessing import StandardScaler

INPUT_COLUMNS = ["N", "P", "F0"]
RA_TARGET = "Ra"
GYLIS_TARGET = "Gylis"

# Fixed ARD Matern model for Ra
RA_FEATURE_MODE = "scale_all"
RA_USE_LOG_TARGET = True
RA_PARAMS = {
    "c_value": 0.42638329674243786,
    "length_scale_n": 3.684816348935756,
    "length_scale_p": 8.825687203522195,
    "length_scale_f0": 100.0,
    "noise_level": 1e-08,
    "gpr_alpha": 4.556011373225105e-12,
    "matern_nu": 2.5,
}

# Fixed ARD Matern model for Gylis
GYLIS_FEATURE_MODE = "raw"
GYLIS_USE_LOG_TARGET = False
GYLIS_PARAMS = {
    "c_value": 0.001,
    "length_scale_n": 1.57227,
    "length_scale_p": 0.0022466,
    "length_scale_f0": 11.43,
    "noise_level": 1e-08,
    "gpr_alpha": 2.80953e-10,
    "matern_nu": 2.5,
}


def build_preprocessor(mode: str) -> ColumnTransformer:
    if mode == "raw":
        return ColumnTransformer([("raw", "passthrough", INPUT_COLUMNS)], remainder="drop")
    if mode == "scale_all":
        return ColumnTransformer([("scale", StandardScaler(), INPUT_COLUMNS)], remainder="drop")
    raise ValueError(f"Unsupported feature mode: {mode}")


def build_kernel(params: dict) -> ConstantKernel:
    length_scale = np.array(
        [
            float(params["length_scale_n"]),
            float(params["length_scale_p"]),
            float(params["length_scale_f0"]),
        ],
        dtype=float,
    )
    return (
        ConstantKernel(
            constant_value=float(params["c_value"]),
            constant_value_bounds="fixed",
        )
        * Matern(
            length_scale=length_scale,
            length_scale_bounds="fixed",
            nu=float(params["matern_nu"]),
        )
        + WhiteKernel(
            noise_level=float(params["noise_level"]),
            noise_level_bounds="fixed",
        )
    )


def train_gp_model(
    df: pd.DataFrame,
    target_column: str,
    feature_mode: str,
    use_log_target: bool,
    params: dict,
) -> tuple[ColumnTransformer, GaussianProcessRegressor]:
    train_df = df.dropna(subset=INPUT_COLUMNS + [target_column]).reset_index(drop=True)
    if train_df.empty:
        raise ValueError(f"No rows available to train {target_column} model.")

    X_df = train_df[INPUT_COLUMNS].copy()
    y = train_df[target_column].to_numpy(dtype=float)
    if use_log_target and np.any(y <= -1.0):
        raise ValueError(f"Cannot use log1p for {target_column}: found values <= -1.")
    y_model = np.log1p(y) if use_log_target else y

    preprocessor = build_preprocessor(feature_mode)
    X = preprocessor.fit_transform(X_df)

    model = GaussianProcessRegressor(
        kernel=build_kernel(params),
        alpha=float(params["gpr_alpha"]),
        optimizer=None,
        normalize_y=True,
        random_state=0,
    )
    model.fit(X, y_model)
    return preprocessor, model


def predict_target(
    df: pd.DataFrame,
    preprocessor: ColumnTransformer,
    model: GaussianProcessRegressor,
    use_log_target: bool,
) -> tuple[list[float], int]:
    preds: list[float] = []
    skipped = 0

    for _, row in df.iterrows():
        if row[INPUT_COLUMNS].isna().any():
            preds.append(np.nan)
            skipped += 1
            continue

        x_df = pd.DataFrame(
            [{"N": float(row["N"]), "P": float(row["P"]), "F0": float(row["F0"])}],
            columns=INPUT_COLUMNS,
        )
        x = preprocessor.transform(x_df)
        pred_model = float(model.predict(x)[0])
        pred = float(np.expm1(pred_model)) if use_log_target else pred_model
        preds.append(pred)

    return preds, skipped


def write_predictions(file_path: Path, pred_ra: list[float], pred_gylis: list[float]) -> None:
    wb = load_workbook(file_path)
    ws = wb.active

    # Column C (3): predicted Ra
    ws.cell(row=1, column=3, value="Pred_Ra")
    for excel_row, value in enumerate(pred_ra, start=2):
        ws.cell(row=excel_row, column=3, value=None if pd.isna(value) else float(value))

    # Column H (8): predicted Gylis
    ws.cell(row=1, column=8, value="Pred_Gylis")
    for excel_row, value in enumerate(pred_gylis, start=2):
        ws.cell(row=excel_row, column=8, value=None if pd.isna(value) else float(value))

    wb.save(file_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Train fixed ARD Matern(2.5) GP models for Ra and Gylis and write "
            "Pred_Ra to column C and Pred_Gylis to column H."
        )
    )
    parser.add_argument("--file", type=str, default="surikiuoti_duomenys_Nscan_3.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Do not write to Excel.")
    args = parser.parse_args()

    file_path = Path(args.file)
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    df = pd.read_excel(file_path)
    required_columns = INPUT_COLUMNS + [RA_TARGET, GYLIS_TARGET]
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    print("Training fixed Ra model (Matern 2.5 ARD)...")
    ra_preprocessor, ra_model = train_gp_model(
        df=df,
        target_column=RA_TARGET,
        feature_mode=RA_FEATURE_MODE,
        use_log_target=RA_USE_LOG_TARGET,
        params=RA_PARAMS,
    )

    print("Training fixed Gylis model (Matern 2.5 ARD)...")
    gylis_preprocessor, gylis_model = train_gp_model(
        df=df,
        target_column=GYLIS_TARGET,
        feature_mode=GYLIS_FEATURE_MODE,
        use_log_target=GYLIS_USE_LOG_TARGET,
        params=GYLIS_PARAMS,
    )

    pred_ra, skipped_ra = predict_target(
        df=df,
        preprocessor=ra_preprocessor,
        model=ra_model,
        use_log_target=RA_USE_LOG_TARGET,
    )
    pred_gylis, skipped_gylis = predict_target(
        df=df,
        preprocessor=gylis_preprocessor,
        model=gylis_model,
        use_log_target=GYLIS_USE_LOG_TARGET,
    )

    print(f"Rows total: {len(df)}")
    print(f"Rows skipped for Ra prediction (missing inputs): {skipped_ra}")
    print(f"Rows skipped for Gylis prediction (missing inputs): {skipped_gylis}")
    print("Write targets: Pred_Ra -> column C (3), Pred_Gylis -> column H (8)")

    if args.dry_run:
        print("Dry-run: no write performed.")
        return

    write_predictions(file_path=file_path, pred_ra=pred_ra, pred_gylis=pred_gylis)
    print(f"Done. Predictions written to column C and H in {file_path}.")


if __name__ == "__main__":
    main()
